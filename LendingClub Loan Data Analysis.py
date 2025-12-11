#!/usr/bin/env python
# coding: utf-8

# In[108]:


# Upgrading Typing_extensions
pip install --upgrade typing_extensions


# In[104]:


# Installing Libraries
get_ipython().system('pip install catboost')
get_ipython().system('pip install shap')


# In[1]:


#--------------------------#
### Importing Libraries ### 
#--------------------------#
import pandas as pd
import numpy as np
import pyodbc
import matplotlib.pyplot as plt
import seaborn as sns
import shap
shap.initjs()

from catboost import CatBoostClassifier
from catboost import Pool

from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report, roc_auc_score
from sklearn.metrics import roc_curve
from sklearn.metrics import roc_auc_score, classification_report
from sklearn.utils import resample
from sklearn.calibration import calibration_curve
from sklearn.calibration import CalibratedClassifierCV

from collections import Counter


# In[2]:


# Importing the cleaned LendingClub loan data directly from SQL

conn = pyodbc.connect("Driver={SQL Server};Server=LAPTOP-H5FPNS0H\SQLEXPRESS;Database=Portfolio Project;Trusted_Connection=yes;")
df = pd.read_sql("SELECT * FROM LendingClub_loan_data", conn)


# In[3]:


# Looking at Data

df.head()


# In[4]:


# Checking for the % of missing values in each column for a better understanding

missing_pct = (df.isna().mean() * 100).round(3).sort_values(ascending=False)
print("\nMissing % (top 20):\n", missing_pct.head(20))


# In[5]:


# Dropping columns with highest missing values which cant be used for ML

df.drop(columns=['mths_since_last_record', 'mths_since_last_major_derog'], inplace=True, errors='ignore')


# In[6]:


# Checking whether the data type changed while importing the data from SQL

df.info()
df.describe(include='all', datetime_is_numeric=True)


# In[7]:


#-------------------------------------------------------------------------------#
###  Converting the columns to their correct data type and cleaning the data ###
#-------------------------------------------------------------------------------#


# In[8]:


# Converting all the date columns to datetime data type 

df['issue_d'] = pd.to_datetime(df['issue_d'])
df['earliest_cr_line'] = pd.to_datetime(df['earliest_cr_line'])


# In[9]:


df['issue_d'].head()


# In[10]:


#-----------------------------------#
###  Extracting numeric value ###
#----------------------------- -----#


# In[11]:


# Separating 'months' from the numerical value from term column and converting it to INT 

df['term'] = df['term'].astype(str).str.extract('(\d+)').astype(int)

df['term'].dtype
df['term'].unique()


# In[12]:


# Extracting the numbers from Emp Length

df['emp_length'] = df['emp_length'].replace({
    '10+ years': 10,
    '9 years': 9,
    '8 years': 8,
    '7 years': 7,
    '6 years': 6,
    '5 years': 5,
    '4 years': 4,
    '3 years': 3,
    '2 years': 2,
    '1 year': 1,
    '< 1 year': 0,
    'n/a': None
}).astype('float')


# In[13]:


df['emp_length'].dtype
df['emp_length'].unique()


# In[14]:


# Extracting the first 3 numbers from the zip code column

df['zip_updated'] = df['zip_code'].astype(str).str[:3]


# In[15]:


df['zip_updated'].head()


# In[16]:


# Dropping the zip_code column
df = df.drop(columns=["zip_code"])


# In[17]:


# Extracting year and month into from the date columns into new columns

df['issue_year'] = df['issue_d'].dt.year
df['issue_month'] = df['issue_d'].dt.month

df['earliest_cr_year'] = df['earliest_cr_line'].dt.year
df['earliest_cr_month'] = df['earliest_cr_line'].dt.month


# In[18]:


# Replace NaT in year/month columns with np.nan (CatBoost can handle np.nan)
date_cols = ['issue_year', 'issue_month', 'earliest_cr_year', 'earliest_cr_month']

for col in date_cols:
    df[col] = df[col].astype('float')   # ensure numeric dtype
    df[col] = df[col].replace({pd.NaT: np.nan})


# In[19]:


#Filling the missing values in Earliest_cr_year and month to -1 to avoid interference when we run the model. 
#The model will treat '-1' as a separate category bin. This avoids creating artificial dates and data leakage 
df['earliest_cr_year'] = df['earliest_cr_year'].fillna(-1).astype(int)
df['earliest_cr_month'] = df['earliest_cr_month'].fillna(-1).astype(int)


# In[20]:


#Checking if we have missing values
df.isna().sum()[['earliest_cr_year','earliest_cr_month']]


# In[21]:


#Dropping the original date columns as there's missing values in them and would interfere when we run the model
df = df.drop(columns=["issue_d", "earliest_cr_line"])


# In[22]:


#---------------------------------#
### Checking for Outliers -  ###
#---------------------------------#


# In[23]:


# Choosing numeric columns to cap 

numeric_to_cap = ['loan_amnt','funded_amnt','installment','int_rate','annual_inc',
    'revol_bal','dti','revol_util','tot_coll_amt','tot_cur_bal','total_rev_hi_lim']


# In[24]:


# Saving a copy of the dataset before capping the amount columns
df_before = df[numeric_to_cap].copy()


# In[25]:


#Visualising the outliers in the amount columns before capping

plt.figure(figsize=(14, 6))
df_before.boxplot()
plt.xticks(rotation=45)
plt.title("Boxplot of Numeric Columns BEFORE Outlier Capping")
plt.show()


# In[26]:


for col in numeric_to_cap:
    plt.figure(figsize=(6, 4))
    plt.boxplot(df_before[col].dropna())
    plt.title(f"Outliers in {col} (Before Capping)")
    plt.ylabel(col)
    plt.show()


# In[27]:


# Using winsorization by IQR for amount columns

def cap_outliers_iqr(series, lower_q=0.25, upper_q=0.75, factor=1.5):
    q1 = series.quantile(lower_q)
    q3 = series.quantile(upper_q)
    iqr = q3 - q1
    lower = q1 - factor * iqr
    upper = q3 + factor * iqr
    return series.clip(lower=lower, upper=upper)

numeric_to_cap = [c for c in numeric_to_cap if c in df.columns]

for c in numeric_to_cap:
    df[c] = pd.to_numeric(df[c], errors='coerce')
    df[c] = cap_outliers_iqr(df[c])
    


# In[28]:


# Comparing before vs After capping of amount columns

df_after = df[numeric_to_cap]

comparison = pd.DataFrame({
    'before_min': df_before.min(),
    'after_min':  df_after.min(),
    'before_max': df_before.max(),
    'after_max':  df_after.max()
})

comparison


# In[29]:


# To get all rows where capping actually happened:

changes = df_before[col] != df[col]
df_before.loc[changes, col].head(), df.loc[changes, col].head()


# In[30]:


# Visually comparing before vs After capping of amount columns (for each column)

df_after = df[numeric_to_cap]

for col in numeric_to_cap:
    plt.figure(figsize=(10, 4))

    # Before
    plt.subplot(1, 2, 1)
    plt.boxplot(df_before[col].dropna())
    plt.title(f"{col} - Before")

    # After
    plt.subplot(1, 2, 2)
    plt.boxplot(df_after[col].dropna())
    plt.title(f"{col} - After")

    plt.tight_layout()
    plt.show()



# In[31]:


# Visualising amount columns after capping

plt.figure(figsize=(14, 6))
df_after.boxplot()
plt.xticks(rotation=45)
plt.title("Boxplot of Numeric Columns After Outlier Capping")
plt.show()


# In[32]:


# From the above graphs and table we can see that 'tot_coll_amt' has mostly 0 values along with the missing values
# and there's a major variation between the 0 values and the other amounts which show up as outliers. Hence we will
# be dropping the column as it is unusable for our modelling.


# In[33]:


# Dropping the 'tot_coll_amt'
df.drop('tot_coll_amt', axis=1, inplace=True)


# In[34]:


# Detecting outliers in all the count columns

count_cols = ['delinq_2yrs', 'inq_last_6mths', 'open_acc', 'pub_rec',
    'total_acc', 'acc_now_delinq', 'collections_12_mths_ex_med']

def detect_count_outliers(df, count_cols):
    outlier_summary = {}

    for col in count_cols:
        if col not in df.columns:
            continue
        
        series = df[col].dropna()

        # Summary statistics
        summary = {
            'min': series.min(),
            'max': series.max(),
            'mean': series.mean(),
            'median': series.median(),
            'p95': series.quantile(0.95),
            'p99': series.quantile(0.99),
            'num_unique': series.nunique(),
            'value_counts_top10': series.value_counts().head(10).to_dict()
        }

        # Identify "rare" high values
        threshold = series.quantile(0.99)
        rare_high_outliers = df[df[col] > threshold][col]

        summary['num_rare_high'] = rare_high_outliers.shape[0]
        summary['rare_high_values'] = rare_high_outliers.unique().tolist()

        outlier_summary[col] = summary

    return outlier_summary

# Run it
count_outlier_report = detect_count_outliers(df, count_cols)
count_outlier_report


# In[35]:


# Since none of these count fields above show any problematic outliers we do not need to cap them.
# They are naturally skewed, but that’s expected for credit-related count data.


# In[36]:


# Visualising the outliers of all the count columns

# Keeping only columns that exist in the dataframe
count_cols = [c for c in count_cols if c in df.columns]

# Create layout for subplots
n_cols = 3
n_rows = int(np.ceil(len(count_cols) / n_cols))

plt.figure(figsize=(15, 5 * n_rows))

for i, col in enumerate(count_cols, 1):
    plt.subplot(n_rows, n_cols, i)
    df[col].plot(kind='box')
    plt.title(f"Boxplot: {col}")
    plt.ylabel("Value")

plt.tight_layout()
plt.show()


# In[37]:


#---------------------------------------------------------------#
### Determining if our numeric data is linear or non linear ###
#---------------------------------------------------------------#


# In[38]:


# Cleaning/Mapping our target column - 'Loan_status'

# Mapping each option to good or bad status(0/1)
mapping = {"fully paid": 0,
    "current": 0,
    "in grace period": 0,           # borderline, will be treating as “not yet default”
    "late (16-30 days)": 1,
    "late (31-120 days)": 1,
    "charged off": 1,
    "default": 1,
    "does not meet the credit policy status:fully paid": 0,
    "does not meet the credit policy status:charged off": 1 }

# Mapping to binary
df['loan_default_flag'] = (
    df['loan_status']
    .str.lower()
    .str.replace(".", "", regex=False)
    .str.split(" status: ").str[-1]   # removing "does not meet credit policy..." wrapper
    .map(mapping)
)
               
               
# Preview
df[['loan_status', 'loan_default_flag']].head(20)


# In[39]:


#Checking if we have any missing values in the loan_default_flag column which is our target column 
df['loan_default_flag'].isna().sum()


# In[40]:


#Dropping the loan_status column since now we have the cleaned binary version. We drop this column as the
# CatBoost model may leak target information.
df = df.drop(columns=["loan_status"], errors="ignore")


# In[41]:


df.info()


# In[42]:


# Selecting numeric columns
numeric_cols_1 = ['annual_inc','revol_util','tot_cur_bal','total_rev_hi_lim',
    'delinq_2yrs','inq_last_6mths','open_acc','pub_rec','total_acc',
    'collections_12_mths_ex_med','acc_now_delinq']


# In[43]:


#Checking if our numeric data is linear or not through scatter plot

n_cols = 3                                # Number of plots per row
n_rows = int(np.ceil(len(numeric_cols_1) / n_cols))

plt.figure(figsize=(18, 5 * n_rows))

for i, col in enumerate(numeric_cols_1, 1):
    plt.subplot(n_rows, n_cols, i)
    sns.scatterplot(
        data=df,
        x=col,
        y='loan_default_flag',
        hue='loan_default_flag',
        alpha=0.3,
        legend=False
    )
    plt.title(f"{col} vs loan_default_flag")
    plt.xlabel(col)
    plt.ylabel("loan_default_flag")

plt.tight_layout()
plt.show()


# In[44]:


#---------------------------------------------------------------------#
### Calculating new variables for our ML model to perform better ###
#---------------------------------------------------------------------#


# In[45]:


# Calculating Credit Utilization
df["credit_util_ratio"] = df["revol_bal"] / (df["total_acc"] + 1)

df["revol_util_clean"] = df["revol_util"] / 100


# In[46]:


# Calculating Income to Loan 
df["income_to_loan"] = df["annual_inc"] / (df["loan_amnt"] + 1)


# In[47]:


# Calculating loan Funded Ratio (Shows how much of loan was funded)
df["funded_ratio"] = df["funded_amnt"] / (df["loan_amnt"] + 1)


# In[48]:


# Creating DTI Buckets
#DTI is rarely linear. Bucketing improves signal.
df["dti_bucket"] = pd.cut(df["dti"],
    bins=[-1, 10, 20, 30, 40, 50, 1000],
    labels=["0-10", "11-20", "21-30", "31-40", "41-50", "50+"])


# In[49]:


# Calculating Numeric Interactions
df["dti_x_intrate"] = df["dti"] * df["int_rate"]
df["loan_x_grade"] = df["loan_amnt"] * df["grade"].astype("category").cat.codes


# In[50]:


# Useful One-hot Flags
df["purpose_small_business"] = (df["purpose"] == "small_business").astype(int)


# In[51]:


#-------------------------------#
### Using CatBoost ML Model ###
#-------------------------------#


# In[58]:


# ------------------------------------------
# 1. Selecting target + features
# ------------------------------------------
target = "loan_default_flag"

features = [c for c in df.columns if c != target]


# In[59]:


# ------------------------------------------
# 2. Identifying categorical columns
#    CatBoost handles categorical encodings internally
# ------------------------------------------
cat_features = ['term', 'grade', 'sub_grade', 'emp_length', 'home_ownership',
    'verification_status', 'purpose', 'addr_state', 'application_type', 'zip_updated', 'dti_bucket']


# In[60]:


#Ensuring proper data type
for c in cat_features:
    df[c] = df[c].astype(str)


# In[61]:


#-----------------------------#
### Train/Val/Test split ###
#-----------------------------#

X_train, X_temp, y_train, y_temp = train_test_split(
    df[features], df[target],
    test_size=0.3,
    random_state=42,
    stratify=df[target]
)

X_val, X_test, y_val, y_test = train_test_split(
    X_temp, y_temp,
    test_size=0.5,
    random_state=42,
    stratify=y_temp
)


# In[62]:


# ------------------------------
# Calculate class weights
# ------------------------------

w0 = (y_train == 0).sum()
w1 = (y_train == 1).sum()

scale = w0 / w1  # weight for minority class 1


# In[66]:


# ------------------------------
# Train CatBoost with class weights
# ------------------------------

model = CatBoostClassifier(
    iterations=2000,
    learning_rate=0.03,
    depth=6,
    eval_metric='AUC',
    loss_function='Logloss',
    class_weights=[1, scale],  # Correct format
    random_seed=42,)

model.fit(X_train, y_train, eval_set=(X_val, y_val), cat_features=cat_features,
    use_best_model=True, verbose=200)


# In[67]:


# -------------------------
# Evaluation
# -------------------------

# Validation scores
y_val_prob = model.predict_proba(X_val)[:, 1]

# Calculating Threshold using Youden’s J
fpr, tpr, thresholds = roc_curve(y_val, y_val_prob)
youden = tpr - fpr
best_threshold = thresholds[np.argmax(youden)]

print("Best threshold:", best_threshold)

y_val_pred = (y_val_prob >= best_threshold).astype(int)
print("Validation AUC:", roc_auc_score(y_val, y_val_prob))


# In[68]:


# Test scores
y_test_prob = model.predict_proba(X_test)[:, 1]
y_test_pred = (y_test_prob >= best_threshold).astype(int)
print("Test AUC:", roc_auc_score(y_test, y_test_prob))
print("\nClassification Report (Test):\n", classification_report(y_test, y_test_pred))


# In[69]:


#Saving Model
model.save_model("model.cbm")


# In[70]:


# Loading the model
model = CatBoostClassifier()
model.load_model("model.cbm")


# In[71]:


# Checking probability distributions
plt.hist(y_test_prob[y_test==0], bins=50, alpha=0.6, label='Class 0')
plt.hist(y_test_prob[y_test==1], bins=50, alpha=0.6, label='Class 1')
plt.legend()
plt.show()


# In[62]:


#-------------------------------#
### Model Debugging Checks ###
#-------------------------------#


# In[63]:


#------------------------#
### Leakage Checks ###


# In[72]:


# Target Leakage From Future Information
future_cols = []

for col in df.columns:
    if "pymnt" in col.lower() or "recover" in col.lower() or "last" in col.lower() or "out_prncp" in col.lower():
        future_cols.append(col)

future_cols


# In[73]:


# Leakage from Target Encoded Columns
[col for col in df.columns if "default" in col.lower() or "risk" in col.lower()]


# In[74]:


# Leakage From Date Columns
[col for col in df.columns if "date" in col.lower() or "_d" in col.lower()]


# In[75]:


#------------------------#
### Imbalance Check ###


# In[76]:


# -------------------------
# 1. CHECK BASIC IMBALANCE
# -------------------------
target = "loan_default_flag"

print("Class distribution:")
print(df[target].value_counts())
print("\nPercentage distribution:")
print(df[target].value_counts(normalize=True) * 100)


# In[77]:


# This is highly imbalance. It causes very poor precision for the minority class. 


# In[78]:


# Plot imbalance
plt.figure(figsize=(6,4))
df[target].value_counts().plot(kind='bar')
plt.title("Target Imbalance")
plt.xlabel("Class")
plt.ylabel("Count")
plt.show()


# In[121]:


# -------------------------
# 2. TRAIN/TEST SPLIT
# -------------------------
features = [c for c in df.columns if c != target]

X_train, X_test, y_train, y_test = train_test_split(df[features],
    df[target],
    test_size=0.2,
    random_state=42,
    stratify=df[target])


# In[122]:


# -------------------------
# 3. CLASS WEIGHTS CALCULATION
# -------------------------
num_neg = (y_train == 0).sum()
num_pos = (y_train == 1).sum()

class_weights = {0: num_pos/num_neg, 1: num_neg/num_pos}
print("\nClass Weights Used:", class_weights)


# In[73]:


# -------------------------
# 4. BASELINE MODEL
# -------------------------
cat_features = X_train.select_dtypes(include=['object']).columns.tolist()

base_model = CatBoostClassifier(
    iterations=300,
    learning_rate=0.05,
    depth=6,
    loss_function='Logloss',
    eval_metric='AUC',
    class_weights=class_weights,
    verbose=0)

base_model.fit(X_train, y_train, eval_set=(X_test, y_test), cat_features=cat_features)


# In[123]:


pred_prob = base_model.predict_proba(X_test)[:,1]
print("\nBaseline AUC:", roc_auc_score(y_test, pred_prob))

threshold = 0.2
pred_class = (pred_prob >= threshold).astype(int)
print("\nClassification Report (Base Model):\n")
print(classification_report(y_test, pred_class))


# In[75]:


# This model used class weights instead of modifying the dataset. Here the AUC is a solid ~0.703. 
# It means CatBoost + class weights is learning useful patterns despite severe imbalance.


# In[93]:


# Saving the model
base_model.save_model("base_model.cbm")


# In[94]:


# Loading the model
base_model = CatBoostClassifier()
base_model.load_model("base_model.cbm")


# In[82]:


# -------------------------
# 5. OVERSAMPLING TEST
# -------------------------
df_train = pd.concat([X_train, y_train], axis=1)

minority = df_train[df_train[target] == 1]
majority = df_train[df_train[target] == 0]


# In[83]:


minority_oversampled = resample(minority, replace=True,
    n_samples=len(majority), random_state=42)

df_oversampled = pd.concat([majority, minority_oversampled])
X_os = df_oversampled[features]
y_os = df_oversampled[target]


# In[79]:


os_model = CatBoostClassifier(
    iterations=300,
    learning_rate=0.05,
    depth=6,
    loss_function='Logloss',
    eval_metric='AUC',
    verbose=0)

os_model.fit(X_os, y_os, eval_set=(X_test, y_test),
    cat_features=cat_features)


# In[80]:


pred_prob_os = os_model.predict_proba(X_test)[:,1]
print("\nOversampled AUC:", roc_auc_score(y_test, pred_prob_os))

pred_os = (pred_prob_os >= threshold).astype(int)
print("\nClassification Report (Oversampled Model):\n")
print(classification_report(y_test, pred_os))


# In[81]:


# This is SMOTE or random oversampling. Oversampling reduced AUC (0.686). 
# This can harm tree-based models like CatBoost


# In[91]:


# Saving the model
os_model.save_model("os_model.cbm")


# In[92]:


# Loading the model
os_model = CatBoostClassifier()
os_model.load_model("os_model.cbm")


# In[85]:


# -------------------------
# 6. UNDERSAMPLING TEST
# -------------------------
minority = df_train[df_train[target] == 1]
majority_downsampled = resample(majority, replace=False,
    n_samples=len(minority), random_state=42)


# In[86]:


df_undersampled = pd.concat([majority_downsampled, minority])
X_us = df_undersampled[features]
y_us = df_undersampled[target]


# In[85]:


us_model = CatBoostClassifier(
    iterations=300,
    learning_rate=0.05,
    depth=6,
    loss_function='Logloss',
    eval_metric='AUC',
    verbose=0)

us_model.fit(X_us, y_us, eval_set=(X_test, y_test),
    cat_features=cat_features)


# In[86]:


pred_prob_us = us_model.predict_proba(X_test)[:,1]
print("\nUndersampled AUC:", roc_auc_score(y_test, pred_prob_us))

pred_us = (pred_prob_us >= threshold).astype(int)
print("\nClassification Report (Undersampled Model):\n")
print(classification_report(y_test, pred_us))


# In[87]:


# This trained on: All minority class samples (1's) and a small random subset of majority class (0's).
# Undersampling produced slightly better AUC (0.7056) than baseline. This means:
# Some majority samples are uninformative and the model benefits from focusing on "hard negatives".
# But undersampling reduces dataset size. This produces the risk of losing information.


# In[89]:


# Saving the model
us_model.save_model("us_model.cbm")


# In[90]:


# Loading the model
us_model = CatBoostClassifier()
us_model.load_model("us_model.cbm")


# In[95]:


# -------------------------
# 7. COMPARE ALL MODELS
# -------------------------
print("\n=======================")
print("AUC SUMMARY")
print("=======================")
print("Baseline (Weighted):", roc_auc_score(y_test, pred_prob))
print("Oversampled:", roc_auc_score(y_test, pred_prob_os))
print("Undersampled:", roc_auc_score(y_test, pred_prob_us))


# In[90]:


# Class-weighted CatBoost model is currently the best clean-performing model so far with AUC 0.703. 
# This will be our final model for debugging. 
# The imbalance is NOT the main reason for the low recall or precision. 
# Our model performs as expected for credit-risk data:
# 1) High recall for minority class (defaults)
# 2) Lower precision
# 3) Strong separation ability (AUC > 0.70)


# In[91]:


#------------------------------#
### Feature Importance ###
#------------------------------#


# In[100]:


# PredictionValueChange Importance

importances = model.get_feature_importance(type="PredictionValuesChange")
fi_df = pd.DataFrame({
    "feature": features,
    "importance": importances
}).sort_values("importance", ascending=False)

print(fi_df.head(20))


# In[101]:


# Plot
plt.figure(figsize=(8, 12))
plt.barh(fi_df["feature"].head(25)[::-1], fi_df["importance"].head(25)[::-1])
plt.title("Top 25 Feature Importance (PredictionValueChange)")
plt.xlabel("Importance Score")
plt.tight_layout()
plt.show()


# In[102]:


# a)From the above Feature importance plot and values we can see that member_id and id are
#   showing importance (5.409, 5.029) almost as strong as annual_inc. This doesn't make sense 
#   as any identifier should have zero relationship with loan default risk. This indicated leakage
#   or pattern memorization. Hence we need to drop both member_id and id columns. 
# b)On the other hand we have many traditional credit features showing high importance such as 
#   Int_rate, annual_inc, sub_grade, grade, dti etc
# c)We should convert earliest_cr_year to credit_age. “Age of credit history” is one of the most powerful
#   predictors in lending risk models. Right now CatBoost only sees the year (e.g., 1998). 
#   It doesn’t understand this as “27 years old”. Converting to an age gives the model an interpretable, normalized feature.
# d)We need to create bins for mths_since_last_delinq because this variable is highly skewed,
#   has many missing values and is non-linear. Binning usually improves performance.


# In[103]:


# Dropping the identifier columns from our dataset
df = df.drop(columns=["member_id", "id"])


# In[104]:


# Creating credit_age from earliest_cr_year
df["credit_age"] = 2025 - df["earliest_cr_year"]


# In[105]:


# Creating bins for mths_since_last_delinq
df["delinq_bin"] = pd.cut(df["mths_since_last_delinq"],
    bins=[-1, 0, 12, 36, 60, 120, 9999],
    labels=["0_months", "1yr", "1-3yr", "3-5yr", "5-10yr", "10yr+"])


# In[106]:


# Dropping mths_since_last_delinq column as it can cause confusion or overlapping in the model
df = df.drop(columns=["mths_since_last_delinq"])


# In[107]:


#-----------------------------------------------------------------------#
### Re-running the CatBoost model after making the necessary changes
#-----------------------------------------------------------------------#


# In[134]:


# ------------------------------------------
# 1. Selecting target + features
# ------------------------------------------
target = "loan_default_flag"

# Excluding the earliest_cr_year from the model to avoid confusion
features = [c for c in df.columns if c not in [target, "earliest_cr_year"]]


# In[135]:


# ------------------------------------------
# 2. Identifying categorical columns
#    CatBoost handles categorical encodings internally
# ------------------------------------------
cat_features = ['term', 'grade', 'sub_grade', 'emp_length', 'home_ownership',
    'verification_status', 'purpose', 'addr_state', 'application_type', 'zip_updated', 
                'dti_bucket', 'delinq_bin']


# In[136]:


#Ensuring proper data type
for c in cat_features:
    df[c] = df[c].astype(str)


# In[137]:


#-----------------------------#
### Train/Val/Test split ###
#-----------------------------#

X_train, X_temp, y_train, y_temp = train_test_split(df[features], df[target],
    test_size=0.3,random_state=42, stratify=df[target])

X_val, X_test, y_val, y_test = train_test_split(X_temp, y_temp,
    test_size=0.5,random_state=42,stratify=y_temp)


# In[138]:


# ------------------------------
# Calculate class weights
# ------------------------------

w0 = (y_train == 0).sum()
w1 = (y_train == 1).sum()

scale = w0 / w1  # weight for minority class 1


# In[139]:


cat_indices = [X_train.columns.get_loc(col) for col in cat_features]


# In[140]:


# ------------------------------
# Train CatBoost with class weights
# ------------------------------

model_1 = CatBoostClassifier(
    iterations=2000,
    learning_rate=0.03,
    depth=6,
    eval_metric='AUC',
    loss_function='Logloss',
    class_weights=[1, scale],  # Correct format
    random_seed=42,)

model_1.fit(X_train, y_train, eval_set=(X_val, y_val), cat_features=cat_indices,
    use_best_model=True, verbose=200)


# In[141]:


# -------------------------
# Evaluation
# -------------------------

# Validation scores
y_val_prob = model_1.predict_proba(X_val)[:, 1]

# Calculating Threshold using Youden’s J
fpr, tpr, thresholds = roc_curve(y_val, y_val_prob)
youden = tpr - fpr
best_threshold = thresholds[np.argmax(youden)]

print("Best threshold:", best_threshold)

y_val_pred = (y_val_prob >= best_threshold).astype(int)
print("Validation AUC:", roc_auc_score(y_val, y_val_prob))


# In[142]:


# Test scores
y_test_prob = model_1.predict_proba(X_test)[:, 1]
y_test_pred = (y_test_prob >= best_threshold).astype(int)
print("Test AUC:", roc_auc_score(y_test, y_test_prob))
print("\nClassification Report (Test):\n", classification_report(y_test, y_test_pred))


# In[145]:


#Saving Model
model_1.save_model("model_1.cbm")


# In[146]:


# Loading model
model_1 = CatBoostClassifier()
model_1.load_model("model_1.cbm")


# In[91]:


#-----------------------------------------------#
### Feature Importance for our Final Model ###
#-----------------------------------------------#


# In[147]:


# PredictionValueChange Importance

importances = model_1.get_feature_importance(type="PredictionValuesChange")
feature_names = X_train.columns

fi_df = pd.DataFrame({
    "feature": feature_names,
    "importance": importances
}).sort_values(by="importance", ascending=False)

print(fi_df)


# In[125]:


# Plot
plt.figure(figsize=(8, 12))
plt.barh(fi_df["feature"].head(25)[::-1], fi_df["importance"].head(25)[::-1])
plt.title("Top 25 Feature Importance (PredictionValueChange)")
plt.xlabel("Importance Score")
plt.tight_layout()
plt.show()


# In[94]:


#------------------------------------------------#
### Calculating SHAP values for the test set ###
#------------------------------------------------#


# In[148]:


# Converting test set into a CatBoost Pool
train_pool = Pool(data=X_train,label=y_train,cat_features=cat_indices)


# In[149]:


# Computing SHAP Values
shap_values = model_1.get_feature_importance(data=train_pool,type="ShapValues")


# In[150]:


# Removing the last column and extracting the actual SHAP values only (n_samples, n_features + 1)
shap_values_for_plot = shap_values[:, :-1]


# In[151]:


# Initialize SHAP explainer for CatBoost
explainer = shap.TreeExplainer(model_1)

# Summary plot
shap.summary_plot(shap_values_for_plot, X_train)


# In[152]:


# SHAP Bar Plot (global average impact)
shap.summary_plot(shap_values_for_plot, X_train, plot_type="bar")


# In[153]:


#SHAP Force Plot for a Single Prediction
i = 10  # example row
shap.force_plot(explainer.expected_value, shap_values_for_plot[i],
    X_train.iloc[i])


# In[154]:


# SHAP Dependence Plot
shap.dependence_plot("dti", shap_values_for_plot, X_train)


# In[155]:


# Checking Calibration
y_test_prob = model_1.predict_proba(X_test)[:, 1]

prob_true, prob_pred = calibration_curve(
    y_test, 
    y_test_prob, 
    n_bins=20,
    strategy="quantile"   # recommended for imbalanced data
)

plt.plot(prob_pred, prob_true, marker='o')
plt.plot([0,1], [0,1], "--")
plt.xlabel("Predicted probability")
plt.ylabel("True probability in each bin")
plt.title("Calibration Curve")
plt.show()


# In[ ]:


### Interpretation of Calibration Curve ###
# The above plot shows:
# Orange dashed line: Perfect calibration - If the model predicts 0.30, then 30% of those cases should truly default.
# Blue curve: model’s actual calibration - How well predicted probabilities match real-world frequencies.
# In the above plot we can see that the curve is well below the diagonal. This means that the model is
# overconfident (overestimates probability of default) For example: When the model predicts 0.70 probability of default
# the actual observed probability is only around 0.30.
# Here the curve rises monotonically (good discrimination) even though probabilities are miscalibrated i.e;
# higher predicted probabilities → higher true default rates. So the model still ranks customers correctly (good AUC).
# This is why our final model has a decent AUC (~0.71) even though calibration is poor. The model is good at 
# ranking risky customers (AUC) but it is not good at giving the true probability of default. It can be used for
# risk ranking, loan approval prioritization and segmentation but raw predicted probabilities shouldn't be used for
# expected loss calculation, pricing and provisioning unless calibrated.


# In[ ]:


#--------------------------------#
### Calibration of the model ###
#--------------------------------#


# In[163]:


# Platt Scaling (Logistic Calibration)
# Fitting the Calibrated Model
calibrated_model = CalibratedClassifierCV(model_1, method='sigmoid', cv='prefit')
calibrated_model.fit(X_val, y_val)


# In[164]:


#Plotting the Sigmoid calibration curve

y_test_prob_cal = calibrated_model.predict_proba(X_test)[:,1]
prob_true, prob_pred = calibration_curve(y_test, y_test_prob_cal, n_bins=20)

plt.plot(prob_pred, prob_true, marker='o', label="Calibrated Model")
plt.plot([0,1],[0,1],"--", label="Perfect Calibration")
plt.title("Calibration Curve (After Sigmoid Calibration)")
plt.legend()
plt.show()


# In[ ]:


### Interpreting the Calibrated Model ###
# The above calibrated plot shows a very good calibration at low–mid probabilities (0.05–0.35).
# From 0.0 to ~0.35 predicted probability, the blue line closely follows the orange line. This means that
# when the model predicts 10% probability, the true default rate is also ~10%. When it predicts 20%, 
# the true rate is ~20%. When it predicts 30–35%, the true rate is close as well. This is strong, reliable 
# calibration for the probability range. We can see a slight under-calibration around 0.4–0.45. This means 
# that when the model predicts 40–45%, the actual default rate is a bit higher than predicted. 
# But our model is still reasonably close and does not break model reliability. That one point around
# 0.45–0.50 where the curve drops to ~0.20 it is a data sparsity artifact. This can happen for a couple of reasons
# like very few samples in high-probability bins or single fit of logistic curve which is less flexible.
# This sometimes causes high-end flattening.
# Overall Interpretation :- 
# The sigmoid-calibrated model is well calibrated across the probability range where most predictions lie (0–0.35).
# It slightly underestimates risk around 40–45%, and the final point at 50% is unstable due to very small bin size.


# In[165]:


# Comparing before vs. after calibration in one plot
# BEFORE calibration
y_test_prob = model_1.predict_proba(X_test)[:, 1]
prob_true_raw, prob_pred_raw = calibration_curve(y_test, y_test_prob, n_bins=20)

# AFTER calibration
y_test_calib_prob = calibrated_model.predict_proba(X_test)[:, 1]
prob_true_cal, prob_pred_cal = calibration_curve(y_test, y_test_calib_prob, n_bins=20)

plt.figure(figsize=(7,5))

plt.plot(prob_pred_raw, prob_true_raw, marker='.', label='Before Calibration')
plt.plot(prob_pred_cal, prob_true_cal, marker='o', label='After Calibration')
plt.plot([0,1], [0,1], "--", label='Perfect calibration')

plt.title("Calibration Curve Comparison")
plt.xlabel("Predicted probability")
plt.ylabel("True probability")
plt.legend()
plt.show()


# In[166]:


# Getting predicted values from our trained data

# Getting predicted classes using your selected threshold
y_test_pred = (y_test_prob >= best_threshold).astype(int)
y_test_pred[:10]


# In[167]:


# Creating a comparison table of real vs predicted

pred_df = pd.DataFrame({
    "actual": y_test,
    "predicted_label": y_test_pred,
    "predicted_probability": y_test_prob
})
pred_df.head()


# In[168]:


# Saving predictions to CSV
pred_df.to_csv("model_predictions.csv", index=False)


# In[169]:


# Seeing which rows were missing & how our Model predicted them

# Inspecting rows that originally had missing values:
rows_with_missing = df[df.isnull().any(axis=1)]
rows_with_missing

# Seeing how the model predicted for those rows after preprocessing
# Align rows with missing values to X_test rows
missing_index = rows_with_missing.index.intersection(X_test.index)

pd.DataFrame({
    "actual": y_test.loc[missing_index],
    "pred_probability": model_1.predict_proba(X_test.loc[missing_index])[:,1]
}).head()


# In[ ]:


#----------------------------------------------------------------#
### # Saving the python code and the results in excel file ###
#----------------------------------------------------------------#


# In[3]:


import openpyxl
from openpyxl.drawing.image import Image
import io
import sys

#-------------------------------------------------
# 1. Load your entire script as text
#-------------------------------------------------
script_path = "C:\Users\91988\LendingClub Loan Data Analysis.py"           # your .py file
excel_path  = "project_export.xlsx"

with open(script_path, "r") as f:
    code_text = f.read()

#-------------------------------------------------
# 2. Capture console outputs
#-------------------------------------------------
buffer = io.StringIO()
sys.stdout = buffer

# Put here the code that prints model outputs
# Example:
print("Validation AUC:", roc_auc_score(y_val, y_val_prob))
print("Best Threshold:", best_threshold)
print("Test AUC:", roc_auc_score(y_test, y_test_prob))
print("Classification Report (Test):\n", classification_report(y_test, y_test_pred))
print("Confusion Matrix:\n", confusion_matrix(y_test, y_test_pred))
prob_true, prob_pred = calibration_curve(y_test, y_test_prob, n_bins=20)
print("Calibration Curve (true vs pred):")
print(list(zip(prob_pred, prob_true)))
print("Feature Importance:\n", fi_df)
print("SHAP values calculated.")

sys.stdout = sys.__stdout__
output_text = buffer.getvalue()

#-------------------------------------------------
# 3. Create Excel workbook
#-------------------------------------------------
wb = openpyxl.Workbook()

# Sheet 1 – Code
ws_code = wb.active
ws_code.title = "Python Code"
ws_code["A1"] = code_text

# Sheet 2 – Outputs
ws_out = wb.create_sheet("Model Outputs")
ws_out["A1"] = output_text

#-------------------------------------------------
# 4. Insert plots
#-------------------------------------------------
# Example plot
plt.figure()
plt.plot([0,1],[0,1])
plt.title("Sample Plot")

# Save plot to buffer
img_buf = io.BytesIO()
plt.savefig(img_buf, format='png')
img_buf.seek(0)

# Insert into Excel
ws_plot = wb.create_sheet("Charts")
img = Image(img_buf)
ws_plot.add_image(img, "A1")

#-------------------------------------------------
# 5. Save Excel file
#-------------------------------------------------
wb.save(excel_path)

print("Export complete:", excel_path)


# In[4]:


import openpyxl
from openpyxl.drawing.image import Image
import subprocess
import matplotlib.pyplot as plt
import sys

# -------------------------------------------------------
# 1. PATHS
# -------------------------------------------------------

# Full path to your python script
script_path = r"C:\Users\91988\LendingClub Loan Data Analysis.py"

# Output Excel file
excel_path = r"C:\Users\91988\OneDrive\Desktop\Data Analyst Portfolio\Python Project\Final_Project_Output.xlsx"

# Folder where your plots are saved
plot_folder = r"C:\Users\91988\OneDrive\Desktop\Data Analyst Portfolio\Python Project\Plots"


# -------------------------------------------------------
# 2. READ YOUR PYTHON CODE AS TEXT
# -------------------------------------------------------
with open(script_path, "r", encoding="utf-8") as f:
    python_code = f.read()


# -------------------------------------------------------
# 3. RUN THE SCRIPT & CAPTURE OUTPUTS
# -------------------------------------------------------
result = subprocess.run(
    ["python", script_path],
    capture_output=True,
    text=True
)

script_output = result.stdout


# -------------------------------------------------------
# 4. CREATE EXCEL FILE
# -------------------------------------------------------
wb = openpyxl.Workbook()

# ---- Sheet 1: Python Code ----
ws1 = wb.active
ws1.title = "Python Code"
ws1["A1"] = python_code


# ---- Sheet 2: Script Outputs (printed results) ----
ws2 = wb.create_sheet("Model Outputs")
ws2["A1"] = script_output


# -------------------------------------------------------
# 5. INSERT PLOTS INTO SEPARATE SHEETS
#   (Save your plots first using plt.savefig("plot1.png"))
# -------------------------------------------------------

import os

plot_files = [f for f in os.listdir(plot_folder) if f.lower().endswith((".png", ".jpg"))]

for plot in plot_files:
    ws = wb.create_sheet(plot.split(".")[0])  # sheet name = plot file name
    img = Image(os.path.join(plot_folder, plot))
    ws.add_image(img, "A1")


# -------------------------------------------------------
# 6. SAVE EVERYTHING
# -------------------------------------------------------
wb.save(excel_path)

print("Excel export complete! File saved at:")
print(excel_path)


# In[6]:


import os

for root, dirs, files in os.walk("C:\\Users\\91988"):
    for f in files:
        if "LendingClub" in f:
            print(os.path.join(root, f))


# In[ ]:


# ---------------------------
# 9. Save cleaned data
# ---------------------------
#df.to_parquet('LendingClub_loan_data_imputed.parquet', index=False)
#df.to_csv('LendingClub_loan_data_imputed.csv', index=False)
#print("\nSaved cleaned files: LendingClub_loan_data_imputed.parquet and .csv")

